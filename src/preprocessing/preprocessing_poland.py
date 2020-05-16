import logging
from contextlib import closing
from pathlib import Path

import numpy as np
import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook
from xlrd import XLRDError

from src.data.datasets import *


def prepare_family_structure_from_voivodship(data_folder: Path) -> pd.DataFrame:
    """
    Preprocesses the family structure excel for a voivodship from pivoted to melted table for easier further processing.
    """
    if (data_folder / household_family_structure_xlsx.file_name).is_file():
        return pd.read_excel(str(data_folder / household_family_structure_xlsx.file_name))

    headcount_columns = [1, 2, 3, 4, 5, 6, 7]
    try:
        df = pd.read_excel(str(data_folder / voivodship_cities_household_family_structure_xlsx.file_name),
                           sheet_name=voivodship_cities_household_family_structure_xlsx.sheet_name)
    except XLRDError:  # sheet not found
        df_q = pd.read_excel(str(data_folder / voivodship_cities_household_family_structure_xlsx.file_name),
                             sheet_name='quantities')
        df = df_q.copy()
        for column in headcount_columns:
            df[column] = df[column] / df[column].sum()

        # FIXME: append a sheet instead of replacing it
        with closing(pd.ExcelWriter(str(data_folder / voivodship_cities_household_family_structure_xlsx.file_name),
                                    engine='openpyxl')) as writer:
            df.to_excel(writer, sheet_name=voivodship_cities_household_family_structure_xlsx.sheet_name, index=False)

    df2 = pd.melt(df,
                  id_vars=['family_type', 'relationship', 'house master'],
                  value_vars=[1, 2, 3, 4, 5, 6, 7], var_name='household_headcount',
                  value_name='probability_within_headcount').rename(columns={'house master': 'house_master'})

    df2.to_excel(str(data_folder / household_family_structure_xlsx.file_name), index=False)
    return df2


def temporary_hack(fcn):
    # FIXME: introducing a hack that is actually going to destroy the probability distribution
    # since it is not possible to have 3 generations in a household where only 2 people live
    # but there is no distribution given for that

    # the input data has to be replaced with a proper distribution that also depends on headcount to have these values
    # correct
    def inner(df, headcount, family_type, relationship, house_master):
        out_df = fcn(df, headcount, family_type, relationship, house_master)

        if len(out_df[out_df['nb_generations'] <= headcount]) > 0:
            out_df = out_df[out_df['nb_generations'] <= headcount]
            out_df['probability'] /= out_df['probability'].sum()
        return out_df

    return inner


@temporary_hack
def draw_generation_configuration_for_household(df, headcount, family_type, relationship, house_master):
    """
    Given a headcount, family type (0,1,2,3), relationship between families
    (if applicable) and who the housemaster is (in multi-family households), this method returns all matching
    households in the df dataframe.
    """

    if house_master not in (np.nan, '', None) and isinstance(house_master, str):
        house_master = house_master.strip()
    if relationship not in (np.nan, '', None) and isinstance(relationship, str):
        relationship = relationship.strip()

    if family_type == 1:
        if house_master not in (np.nan, '', None):
            return df[(df.family_type == family_type) & (df.relationship == relationship)
                      & (df.house_master == house_master)]
        return df[(df.family_type == family_type) & (df.relationship == relationship)]
    if family_type == 2:
        if house_master not in (np.nan, '', None):
            out_df = df[(df.family_type == family_type) & (df.relationship == relationship)
                        & (df.house_master == house_master)]
            if len(out_df) > 0:
                return out_df
        if relationship not in (np.nan, '', None):
            out_df = df[(df.family_type == family_type) & (df.relationship == relationship)]
            if len(out_df) > 0:
                return out_df
        return df[(df.family_type == family_type)]
    if family_type == 3:
        return df[(df.family_type == family_type)]
    if family_type == 0:
        if headcount == 1:
            return df[(df.family_type == family_type) & (df.relationship == 'Jednoosobowe')]
        return df[(df.family_type == family_type) & (df.relationship == 'Wieloosobowe')]
    raise ValueError(f'Unknown family type {family_type}')


def _sanitize_households_count(households_count_df, population_size):
    old_population = (households_count_df['nb_of_people_in_household'] * households_count_df['nb_of_households']).sum()
    households_count_df['nb_of_households'] *= (population_size / old_population)
    households_count_df['nb_of_households'] = households_count_df['nb_of_households'].apply(np.ceil).astype(int)
    assert (households_count_df['nb_of_people_in_household'] * households_count_df['nb_of_households']).sum() \
           >= population_size


def _filter_family_structures_for_household(family_structure_df, hc_row):
    fs_df = family_structure_df[family_structure_df.household_headcount == hc_row.nb_of_people_in_household].copy()
    fs_df['total'] = (np.round(fs_df.probability_within_headcount * hc_row.nb_of_households)).astype(int)

    difference_due_to_rounding = hc_row.nb_of_households - fs_df['total'].sum()
    if difference_due_to_rounding != 0:
        logging.info(f'Difference due to numeric error {difference_due_to_rounding} - assigning randomly')
        try:
            fs_df.loc[np.random.choice(fs_df[fs_df.total > 0].index.tolist()), 'total'] += difference_due_to_rounding
        except ValueError:
            fs_df.loc[np.random.choice(fs_df.index.tolist()), 'total'] += difference_due_to_rounding

    return fs_df


def get_generations_configuration_df(data_folder: Path, xlsx_file: XlsxFile) -> pd.DataFrame:
    generations_configuration_df = pd.read_excel(str(data_folder / xlsx_file.file_name),
                                                 sheet_name=xlsx_file.sheet_name)
    generations_configuration_df['nb_generations'] = generations_configuration_df['young'] \
                                                     + generations_configuration_df['middle'] \
                                                     + generations_configuration_df['elderly']
    return generations_configuration_df


def generate_household_indices(data_folder: Path, output_folder: Path, population_size: int) -> pd.DataFrame:
    """Generates and saves to an excel file a dataframe of households. Each household consists of:
     * an index,
     * headcount,
     * family_type (0,1,2,3 - number of families in the household)
     * relationship - between families, if more than one lives in the household
     * house_master - in 2 and 3 families households, which family does the housemaster belong to
     * family_structure_regex - auxiliary description of a household
     * young - flag whether people younger than 30 years old live in a household
     * middle - flag, whether people between 30 and 59 inclusive live in a household
     * elderly - flag, whether people older than 59 live in a household

    :param data_folder: data folder
    :param output_folder: where to save an output file
    :param population_size: size of a population needing accommodation
    :return:
    """
    household_headcount = []
    family_type = []
    relationship = []
    house_master = []
    young = []
    middle = []
    elderly = []

    family_structure_df = pd.read_excel(str(data_folder / household_family_structure_xlsx.file_name),
                                        sheet_name=household_family_structure_xlsx.sheet_name)
    households_count_df = pd.read_excel(str(data_folder / households_count_xlsx.file_name),
                                        sheet_name=households_count_xlsx.sheet_name)
    generations_configuration_df = get_generations_configuration_df(data_folder, generations_configuration_xlsx)

    _sanitize_households_count(households_count_df, population_size)

    for i, hc_row in tqdm(households_count_df.iterrows(), total=len(households_count_df.index)):
        # family structure given this headcount
        fs_df = _filter_family_structures_for_household(family_structure_df, hc_row)

        for j, row in fs_df[fs_df.total > 0].iterrows():
            household_headcount.extend([row.household_headcount] * row.total)
            family_type.extend([row.family_type] * row.total)
            relationship.extend([row.relationship] * row.total)
            house_master.extend([row.house_master] * row.total)

            gc_df = draw_generation_configuration_for_household(generations_configuration_df,
                                                                row.household_headcount,
                                                                row.family_type,
                                                                row.relationship,
                                                                row.house_master)
            gc_idx = np.random.choice(gc_df.index.tolist(), p=gc_df.probability, size=row.total)
            young.extend(gc_df.loc[gc_idx, 'young'])
            middle.extend(gc_df.loc[gc_idx, 'middle'])
            elderly.extend(gc_df.loc[gc_idx, 'elderly'])

    household_df = pd.DataFrame(data=dict(household_index=list(range(len(household_headcount))),
                                          household_headcount=household_headcount,
                                          family_type=family_type,
                                          relationship=relationship,
                                          house_master=house_master,
                                          young=young, middle=middle, elderly=elderly))

    household_df.to_feather(str(output_folder / output_households_basic_feather.file_name))
    return household_df


def generate_generations_configuration(data_folder: Path) -> pd.DataFrame:
    """
    This function does the preprocessing of Census data for age generations living together in households:

    Generations - a table that contains probability of living together. In the original table there are seven columns:
    * young alone -> cat1
    * middle-aged alone -> cat2
    * elderly alone -> cat3
    * young and middle-aged together -> cat4
    * young and elderly together -> cat5
    * middle-aged and elderly together -> cat6
    * young, middle-aged and elderly together -> cat7
    The function takes occurrences of each category and models them as three boolean columns: young, middle, elderly.

    Additionally, family_type field is changed from descriptive, string form into a number (0, 1, 2, 3) that represents
    the number of families living in a household.
    """
    output_file = data_folder / generations_configuration_xlsx.file_name
    if output_file.is_file():
        return pd.read_excel(str(output_file), sheet_name=generations_configuration_xlsx.sheet_name)

    voivodship_workbook_path = str(data_folder / voivodship_cities_generations_configuration_xlsx.file_name)
    v_config_df = pd.read_excel(voivodship_workbook_path, sheet_name='preprocessed', header=[0, 1])
    melted = pd.melt(v_config_df, id_vars=[('Unnamed: 0_level_0', 'family_type'),
                                           ('Unnamed: 1_level_0', 'relationship'),
                                           ('Unnamed: 2_level_0', 'house_master')],
                     var_name=['unit', 'category'],
                     value_name='total')
    melted = melted.rename(columns={('Unnamed: 0_level_0', 'family_type'): 'family_type',
                                    ('Unnamed: 1_level_0', 'relationship'): 'relationship',
                                    ('Unnamed: 2_level_0', 'house_master'): 'house_master'})
    melted['young'] = melted.category.isin(['cat1', 'cat4', 'cat5', 'cat7']).astype(int)
    melted['middle'] = melted.category.isin(['cat2', 'cat4', 'cat6', 'cat7']).astype(int)
    melted['elderly'] = melted.category.isin(['cat3', 'cat5', 'cat6', 'cat7']).astype(int)
    melted = melted[melted.category != 'total']
    melted = melted.drop(columns=['category'])
    melted['relationship'] = melted['relationship'].fillna('N/A')
    melted['house_master'] = melted['house_master'].fillna('N/A')
    pivoted = pd.pivot_table(melted, columns=['unit'], values='total',
                             index=['family_type', 'relationship', 'house_master', 'young', 'middle', 'elderly'],
                             aggfunc='first').reset_index()
    pivoted.households = pd.to_numeric(pivoted.households, errors='coerce')
    pivoted.people = pd.to_numeric(pivoted.people, errors='coerce')
    pivoted = pivoted.fillna(0)

    pivoted.loc[pivoted['family_type'] == 'Jednorodzinne', 'family_type'] = 1
    pivoted.loc[pivoted['family_type'] == 'Dwurodzinne', 'family_type'] = 2
    pivoted.loc[pivoted['family_type'] == 'Trzy i więcej rodzinne', 'family_type'] = 3
    pivoted.loc[pivoted['family_type'] == 'Nierodzinne', 'family_type'] = 0

    book = load_workbook(voivodship_workbook_path)

    if voivodship_cities_generations_configuration_xlsx.sheet_name in book.sheetnames:
        del book[voivodship_cities_generations_configuration_xlsx.sheet_name]

    with closing(pd.ExcelWriter(voivodship_workbook_path, engine='openpyxl')) as writer:
        writer.book = book
        pivoted.to_excel(writer, sheet_name=voivodship_cities_generations_configuration_xlsx.sheet_name, index=False)
        writer.save()

    # update with probabilities
    df = pivoted.groupby(by=['family_type', 'relationship', 'house_master'])['households'].sum().reset_index() \
        .rename(columns={'households': 'total'})
    pivoted = pivoted.merge(df, how='left', on=['family_type', 'relationship', 'house_master'])
    pivoted['probability'] = pivoted['households'] / pivoted['total']

    pivoted.to_excel(str(output_file), sheet_name=generations_configuration_xlsx.sheet_name, index=False)
    return pivoted


def voivodship_to_symbol(voivodship: str):
    voivodship = voivodship.lower()
    voivodship_dict = {
        'podlaskie': 'B',
        'kujawsko-pomorskie': 'C',
        'dolnośląskie': 'D',
        'łódzkie': 'E',
        'lubuskie': 'F',
        'pomorskie': 'G',
        'małopolskie': 'K',
        'lubelskie': 'L',
        'warmińsko-mazurskie': 'N',
        'opolskie': 'O',
        'wielkopolskie': 'P',
        'podkarpackie': 'R',
        'śląskie': 'S',
        'świętokrzyskie': 'T',
        'mazowieckie': 'W',
        'zachodniopomorskie': 'Z'}
    return voivodship_dict.get(voivodship)

    if __name__ == '__main__':
        project_dir = Path(__file__).resolve().parents[2]
    city_folder = project_dir / 'data' / 'processed' / 'poland' / 'WW'
    prepare_family_structure_from_voivodship(city_folder)
    generate_generations_configuration(city_folder)
